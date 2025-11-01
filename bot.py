import os
import threading
import logging
import pandas as pd
from dotenv import load_dotenv
from flask import Flask
from telegram import Update, KeyboardButton, ReplyKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes
from openpyxl import load_workbook

# -----------------------------
# تنظیمات لاگ
# -----------------------------
logging.basicConfig(
    format="%(asctime)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# -----------------------------
# بارگذاری متغیرهای محیطی
# -----------------------------
load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")
if not BOT_TOKEN:
    logger.error("❌ BOT_TOKEN not found in environment.")
    raise SystemExit("BOT_TOKEN required")

# -----------------------------
# فایل‌ها (روی ریپوزیتوری GitHub شما باشند)
# -----------------------------
FOC_FILE = "FOC.xlsx"
LIGA_FILE = "Rliga 140408 - TG.xlsx"

# -----------------------------
# بارگذاری داده‌ها (مقاوم در برابر خطا)
# -----------------------------
try:
    # sheet 0: mapping طرح‌ها (شماره طرح، عنوان طرح، TableName)
    df_plans = pd.read_excel(FOC_FILE, sheet_name=0)
    required_cols = ["شماره طرح", "عنوان طرح", "TableName"]
    for c in required_cols:
        if c not in df_plans.columns:
            raise ValueError(f"ستون مورد انتظار '{c}' در شیت 0 فایل FOC پیدا نشد.")
    # نگاشت عنوان -> شماره و عنوان -> TableName
    title_to_number = dict(zip(df_plans["عنوان طرح"].astype(str).str.strip(), df_plans["شماره طرح"]))
    title_to_table = dict(zip(df_plans["عنوان طرح"].astype(str).str.strip(), df_plans["TableName"].astype(str).str.strip()))

    # sheet 1: سوالات مربوط به هر طرح (ستون 'شماره طرح' و یک یا چند ستون سوال)
    df_questions = pd.read_excel(FOC_FILE, sheet_name=1)
    if "شماره طرح" not in df_questions.columns:
        raise ValueError("ستون 'شماره طرح' در شیت 1 فایل FOC موجود نیست.")
    # ستون(های) سوال را همه ستون‌های غیر از 'شماره طرح' در نظر می‌گیریم
    question_columns = [c for c in df_questions.columns if str(c).strip() != "شماره طرح"]
    if not question_columns:
        raise ValueError("ستونی برای سوالات در شیت 1 یافت نشد.")
    # آماده‌سازی mapping: شماره طرح -> لیست سوال‌ها (تمام ستون‌های سوال)
    questions_by_plan = {}
    for plan_no, group in df_questions.groupby("شماره طرح"):
        qlist = []
        for qc in question_columns:
            qlist.extend(group[qc].dropna().astype(str).str.strip().tolist())
        # حذف تکراری و خالی
        qlist = [q for q in pd.Series(qlist).unique().tolist() if q]
        questions_by_plan[plan_no] = qlist

    # سوالات اولیه: از همه ستون‌های سوال در شیت1 می‌آوریم (unique)
    initial_questions = []
    for qc in question_columns:
        initial_questions.extend(df_questions[qc].dropna().astype(str).str.strip().tolist())
    # حذف تکرار و خالی
    initial_questions = [q for q in pd.Series(initial_questions).unique().tolist() if q]

    logger.info("✅ فایل‌های FOC با موفقیت بارگذاری شدند.")
except Exception as e:
    logger.exception("❌ خطا در بارگذاری فایل‌های FOC: %s", e)
    raise

# -----------------------------
# توابع کمکی برای کار با اکسل Rliga و جداول
# -----------------------------
def find_table_in_sheet(wb, sheet_name, desired_table_name):
    """
    پیدا کردن جدول (Table) داخل شیت داده‌شده به صورت case-insensitive و trim.
    برمی‌گرداند شیء openpyxl Table یا None.
    """
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    desired = str(desired_table_name).strip().lower()
    for tname, tbl in ws.tables.items():
        if tname.strip().lower() == desired:
            return tbl, ws
    # تلاش دیگر: اگر TableName ممکنه ذخیره نشده باشه، سعی کن exact match با کلیدها یا با فاصله‌ها
    for tname, tbl in ws.tables.items():
        if tname.strip().lower().replace(" ", "") == desired.replace(" ", ""):
            return tbl, ws
    return None

def table_to_dataframe(tbl, ws):
    """
    تبدیل openpyxl table به pandas DataFrame.
    """
    ref = tbl.ref  # مثال: 'A1:F100'
    # ws[ref] به‌شکل tuple از ردیف‌ها برمی‌گرده
    cells = ws[ref]
    header = [c.value for c in cells[0]]
    rows = []
    for r in cells[1:]:
        rows.append([c.value for c in r])
    df = pd.DataFrame(rows, columns=header)
    return df

# -----------------------------
# ربات تلگرام
# -----------------------------
app = ApplicationBuilder().token(BOT_TOKEN).build()

# کیبورد سوالات اولیه
keyboard_initial = [[KeyboardButton(q)] for q in initial_questions]
reply_markup_initial = ReplyKeyboardMarkup(keyboard_initial, one_time_keyboard=True, resize_keyboard=True)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("👋 سلام! لطفاً یک سوال اولیه انتخاب کنید:", reply_markup=reply_markup_initial)
    context.user_data["state"] = "choosing_initial_question"

# تابع کمکی برای ارسال لیست طرح‌ها
def plans_keyboard():
    kb = [[KeyboardButton(t)] for t in title_to_number.keys()]
    return ReplyKeyboardMarkup(kb, one_time_keyboard=True, resize_keyboard=True)

# توابع محاسبه پاسخ (پایه‌ای — ممکنه نیاز به تنظیم با ساختار جدول شما داشته باشه)
def get_top_n(df, metric_col=None, n=5, name_col_candidates=["نام", "نام و نام خانوادگی", "نام خانوادگی", "نام کامل"]):
    # تلاش برای پیدا کردن ستون اسم
    name_col = None
    for c in name_col_candidates:
        if c in df.columns:
            name_col = c
            break
    if metric_col and metric_col in df.columns:
        df_sorted = df.sort_values(by=metric_col, ascending=False)
    else:
        # اگر ستونی برای metric ندیدیم سعی کن بر اساس 'رتبه' یا ستون عددی اول
        if "رتبه" in df.columns:
            df_sorted = df.sort_values(by="رتبه", ascending=True)
        else:
            numeric_cols = df.select_dtypes(include="number").columns.tolist()
            if numeric_cols:
                df_sorted = df.sort_values(by=numeric_cols[0], ascending=False)
            else:
                return []
    if name_col:
        return df_sorted[name_col].head(n).astype(str).tolist()
    else:
        # اگر اسم پیدا نشد، باز کل سطرها رو برمی‌گردونیم (نمایش فیلدهای مهم)
        return df_sorted.head(n).to_dict(orient="records")

def find_rank_by_empid(df, emp_id, emp_col_candidates=["کد پرسنلی", "کد", "emp_id"]):
    emp_col = None
    for c in emp_col_candidates:
        if c in df.columns:
            emp_col = c
            break
    if emp_col is None:
        return None
    row = df[df[emp_col].astype(str) == str(emp_id)]
    if row.empty:
        return None
    # تلاش برای یافتن ستون رتبه
    if "رتبه" in df.columns:
        return row["رتبه"].values[0]
    # اگر رتبه نیست، تلاش برای محاسبه based on numeric sort
    numeric_cols = df.select_dtypes(include="number").columns.tolist()
    if numeric_cols:
        # فرض: بزرگترین مقدار بهترین -> محاسبه رتبه
        col = numeric_cols[0]
        df_sorted = df.sort_values(by=col, ascending=False).reset_index(drop=True)
        pos = df_sorted.index[df_sorted[emp_col].astype(str) == str(emp_id)].tolist()
        if pos:
            return pos[0] + 1
    return None

# handler پیام‌ها
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (update.message.text or "").strip()
    state = context.user_data.get("state", "")
    try:
        # مرحله انتخاب سوال اولیه
        if state == "choosing_initial_question" or not state:
            # نگهدار سوال اولیه انتخاب شده
            context.user_data["initial_question"] = text
            # نمایش لیست طرح‌ها
            await update.message.reply_text("📋 لطفاً طرح مورد نظر خود را انتخاب کنید:", reply_markup=plans_keyboard())
            context.user_data["state"] = "choosing_plan"
            return

        # مرحله انتخاب طرح
        if state == "choosing_plan":
            # تطابق عنوان طرح با نگاشت
            key = None
            for t in title_to_number.keys():
                if t.strip() == text:
                    key = t
                    break
            if key is None:
                await update.message.reply_text("❌ طرح یافت نشد، لطفاً دوباره انتخاب کنید.", reply_markup=plans_keyboard())
                return
            selected_number = title_to_number[key]
            context.user_data["selected_number"] = selected_number
            context.user_data["selected_title"] = key
            context.user_data["selected_table"] = title_to_table.get(key)
            # سوالات مربوط به طرح
            qs = questions_by_plan.get(selected_number, [])
            if not qs:
                await update.message.reply_text("❌ سوالی برای این طرح موجود نیست.")
                # بازگشت به انتخاب طرح
                await update.message.reply_text("📋 لطفاً طرح دیگری انتخاب کنید:", reply_markup=plans_keyboard())
                context.user_data["state"] = "choosing_plan"
                return
            keyboard_qs = [[KeyboardButton(q)] for q in qs]
            await update.message.reply_text("📋 لطفاً سوال خود را انتخاب کنید:", reply_markup=ReplyKeyboardMarkup(keyboard_qs, one_time_keyboard=True, resize_keyboard=True))
            context.user_data["state"] = "choosing_question"
            return

        # مرحله انتخاب سوال
        if state == "choosing_question":
            selected_question = text
            table_name = context.user_data.get("selected_table")
            if not table_name:
                await update.message.reply_text("❌ نام جدول برای این طرح تنظیم نشده.")
                return

            # بارگذاری workbook و تلاش برای پیدا کردن Table در شیت "فروشنده"
            try:
                wb = load_workbook(LIGA_FILE, data_only=True)
            except Exception as e:
                logger.exception("خطا در باز کردن فایل لیگا: %s", e)
                await update.message.reply_text("❌ خطا در باز کردن فایل داده‌ها.")
                return

            found = find_table_in_sheet(wb, "فروشنده", table_name)
            if not found:
                # تلاش fallback: جستجو در همه شیت‌ها برای جداول
                tbl_ws = None
                for sname in wb.sheetnames:
                    tmp = find_table_in_sheet(wb, sname, table_name)
                    if tmp:
                        found = tmp
                        break
            if not found:
                await update.message.reply_text(f"❌ Table با نام '{table_name}' در شیت 'فروشنده' پیدا نشد.")
                return
            tbl, ws = found
            df_table = table_to_dataframe(tbl, ws)

            # اگر سوال نیاز به کد پرسنلی داره (مثلاً 'رتبه من چندمه' یا حاوی 'کد پرسنلی')
            if "کد پرسنلی" in selected_question or "رتبه من" in selected_question or "رتبهٔ من" in selected_question:
                await update.message.reply_text("لطفاً کد پرسنلی خود را وارد کنید:")
                context.user_data["state"] = "waiting_for_id"
                context.user_data["last_question"] = selected_question
                context.user_data["df_table_cached"] = df_table
                return

            # پاسخ‌های رایج:
            # - "نفر اول کیه؟" -> سطر با رتبه 1 یا بزرگترین مقدار در ستون عددی
            # - "5نفر اول چه کسانی هستن؟" -> top5
            # - "فاصله من با نفر اول چند مشتری هستش؟" -> نیازمند کد پرسنلی -> اگر emp id not provided، پیام دهیم.
            q = selected_question
            if q.strip().startswith("نفر اول"):
                # تلاش برای پیدا کردن ستون اسم و رتبه
                if "رتبه" in df_table.columns:
                    row = df_table[df_table["رتبه"] == 1]
                    if not row.empty:
                        # تلاش برای اسم
                        for name_col in ["نام", "نام و نام خانوادگی", "نام کامل", "نام خانوادگی"]:
                            if name_col in df_table.columns:
                                await update.message.reply_text(f"💡 نفر اول: {row.iloc[0][name_col]}")
                                break
                        else:
                            await update.message.reply_text(f"💡 سطر نفر اول:\n{row.iloc[0].to_dict()}")
                        # بازگشت به لیست طرح‌ها
                        await update.message.reply_text("📋 لطفاً طرح دیگری انتخاب کنید:", reply_markup=plans_keyboard())
                        context.user_data["state"] = "choosing_plan"
                        return
                # fallback: ستون عددی بهترین
                top = get_top_n(df_table, n=1)
                if top:
                    await update.message.reply_text(f"💡 نفر اول: {top[0]}")
                else:
                    await update.message.reply_text("❌ نتوانستم نفر اول را محاسبه کنم.")
                await update.message.reply_text("📋 لطفاً طرح دیگری انتخاب کنید:", reply_markup=plans_keyboard())
                context.user_data["state"] = "choosing_plan"
                return

            if "5نفر اول" in q or "۵نفر" in q or q.strip().startswith("5"):
                top5 = get_top_n(df_table, n=5)
                if top5:
                    if isinstance(top5[0], dict):
                        await update.message.reply_text("💡 ۵ نفر اول:\n" + "\n".join(str(r) for r in top5))
                    else:
                        await update.message.reply_text("💡 ۵ نفر اول:\n" + "\n".join(f"{i+1}. {name}" for i, name in enumerate(top5)))
                else:
                    await update.message.reply_text("❌ نتوانستم ۵ نفر اول را محاسبه کنم.")
                await update.message.reply_text("📋 لطفاً طرح دیگری انتخاب کنید:", reply_markup=plans_keyboard())
                context.user_data["state"] = "choosing_plan"
                return

            # fallback: اگر جدول خودش شامل ستون سوال/پاسخ است (همان روشی که شما قبلاً تلاش کرده بودی)
            # جستجو در ستون‌هایی که ممکن است سوال را داشته باشند:
            question_col_candidates = [c for c in df_table.columns if "سوال" in str(c) or "پرسش" in str(c)]
            if question_col_candidates:
                qc = question_col_candidates[0]
                answer_cols = [c for c in df_table.columns if c != qc]
                if answer_cols:
                    row = df_table[df_table[qc].astype(str).str.strip() == q]
                    if not row.empty:
                        ans = row.iloc[0][answer_cols[0]]
                        await update.message.reply_text(f"💡 جواب: {ans}")
                        await update.message.reply_text("📋 لطفاً طرح دیگری انتخاب کنید:", reply_markup=plans_keyboard())
                        context.user_data["state"] = "choosing_plan"
                        return

            # اگر به اینجا رسیدیم، جواب مستقیم نیافتیم
            await update.message.reply_text("❌ نتوانستم جواب سؤال را محاسبه کنم یا الگویِ پشتیبانی‌شده نیست.")
            await update.message.reply_text("📋 لطفاً طرح دیگری انتخاب کنید:", reply_markup=plans_keyboard())
            context.user_data["state"] = "choosing_plan"
            return

        # حالت انتظار برای کد پرسنلی
        if state == "waiting_for_id":
            emp_id = text
            df_table = context.user_data.get("df_table_cached")
            if df_table is None:
                await update.message.reply_text("❌ اطلاعات جدول در حافظه وجود ندارد؛ لطفاً دوباره سوال را انتخاب کنید.")
                context.user_data["state"] = "choosing_plan"
                return
            rank = find_rank_by_empid(df_table, emp_id)
            if rank is None:
                await update.message.reply_text("❌ کد پرسنلی یافت نشد یا ستون‌های مورد نیاز موجود نیست.")
            else:
                await update.message.reply_text(f"💡 رتبه شما: {rank}")
            # بازگرداندن لیست سوالات مربوطه
            selected_number = context.user_data.get("selected_number")
            qs = questions_by_plan.get(selected_number, [])
            keyboard_qs = [[KeyboardButton(q)] for q in qs]
            await update.message.reply_text("📋 لطفاً سوال خود را انتخاب کنید:", reply_markup=ReplyKeyboardMarkup(keyboard_qs, one_time_keyboard=True, resize_keyboard=True))
            context.user_data["state"] = "choosing_question"
            return

        # fallback کلی
        await update.message.reply_text("👋 لطفاً یک سوال اولیه انتخاب کنید:", reply_markup=reply_markup_initial)
        context.user_data["state"] = "choosing_initial_question"

    except Exception as e:
        logger.exception("خطا در پردازش پیام: %s", e)
        await update.message.reply_text(f"❌ خطا در پردازش پیام: {e}")

# افزودن handlerها
app.add_handler(CommandHandler("start", start))
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

# -----------------------------
# Flask healthcheck — حتماً روی پورت $PORT اجرا کنیم (main thread)
# -----------------------------
flask_app = Flask("healthcheck")

@flask_app.route("/")
def home():
    return "Bot is running!"

def start_telegram_polling_in_thread():
    logger.info("Starting telegram polling in background thread...")
    # اجرای polling در thread جدا
    threading.Thread(target=lambda: app.run_polling(), daemon=True).start()

# -----------------------------
# main: راه‌اندازی بات در background و اجرای Flask در main (برای Render)
# -----------------------------
if __name__ == "__main__":
    logger.info("✅ Preparing to start bot + healthcheck...")
    start_telegram_polling_in_thread()
    port = int(os.environ.get("PORT", 10000))
    logger.info("Starting Flask healthcheck on port %s", port)
    # اجرای Flask در main thread (Render این پورت را چک می‌کند)
    flask_app.run(host="0.0.0.0", port=port)
